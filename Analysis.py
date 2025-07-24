# 匯入 Flask 框架及相關模組
from flask import Flask, request, jsonify, render_template, session, send_file
from typer import prompt
from gpt_utils import extract_resolution_suggestion
from gpt_utils import extract_problem_with_custom_prompt
from gpt_utils import analyze_with_ai_builder_then_fallback
from gptChat import run_offline_gpt
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict
from collections import Counter
import hashlib
import subprocess
import sys
from pathlib import Path
import umap
import hdbscan
# 匯入數學運算模組
import math
import json
# 匯入 pandas 用於處理 Excel 資料
import pandas as pd
# 匯入 os 模組處理檔案與路徑
import os
import shutil
# 匯入正則表達式模組
import re
import glob
# 匯入 webbrowser 用於開啟網頁
import webbrowser
import socket
# 匯入 traceback 用於錯誤追蹤
import traceback
# 匯入 Werkzeug 的工具函數確保檔案名稱安全
from werkzeug.utils import secure_filename
# ✅ 匯入語意分析模組
from SmartScoring import is_high_risk, is_escalated, is_multi_user, extract_keywords, recommend_solution, is_actionable_resolution, load_embeddings, load_examples_from_json
# ✅ 預先 encode 一筆資料以加速首次請求
from SmartScoring import bert_model  # 確保你有從 SmartScoring 載入模型
from SmartScoring import extract_cluster_name  # 匯入自定的 cluster 命名函式
from tqdm import tqdm
from sentence_transformers import util
# ✅ 匯入關鍵字抽取模組
from datetime import datetime
from sklearn.cluster import KMeans
import numpy as np
from datetime import datetime
import time
# --- 分群啟用條件（可依資料調整）---
import asyncio
import aiohttp
import math
import requests
import threading
import json
import tempfile
from jsonschema import validate, ValidationError
from datetime import datetime
from collections import Counter
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import traceback
from flask_socketio import SocketIO, emit



POWERAUTOMATE_CLASSIFY_URL = "https://prod-26.southeastasia.logic.azure.com:443/workflows/651f88b4e548481ba38d129c30af1cae/triggers/manual/paths/invoke?api-version=2016-06-01&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=1kjXdbN7QsORisL6sdEA1IRXRef_bstLqZjmRjp9c6E"  # 🔁 改成你自己的分類流程 URL
POWERAUTOMATE_SUMMARY_URL = "https://prod-71.southeastasia.logic.azure.com:443/workflows/d70056c4f2c044b9a297164c9f98d1b6/triggers/manual/paths/invoke?api-version=2016-06-01&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=L_qVgz5s0bLvz20lmI3RsoEmClvbTJfy7v99Ai38Xpw"  # 🔁 改成你自己的摘要流程 URL

KMEANS_MIN_COUNT = 4         # 最少資料筆數
KMEANS_MIN_RANGE = 5.0       # 分數最大最小值差
KMEANS_MIN_STDDEV = 3.0      # 標準差下限

progress_log = ""  # 全域變數，專門用來存進度訊息




start = time.time()
print("🔥 預熱語意模型中...")
bert_model.encode("warmup")  # 預熱一次，避免第一次使用太慢
print(f"✅ 模型預熱完成，用時：{time.time() - start:.2f} 秒")

# 建立 Flask 應用
app = Flask(__name__)
# 設定應用的密鑰，用於 session 加密
app.secret_key = 'gwegweqgt22e'
# 設定 session 儲存方式為檔案系統
app.config['SESSION_TYPE'] = 'filesystem'

# ------------------------------------------------------------------------------
# 設定上傳資料夾與大小限制（10MB）
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 限制檔案大小為 10MB
ALLOWED_EXTENSIONS = {'xlsx'}  # 僅允許上傳 xlsx 檔案


basedir = os.path.abspath(os.path.dirname(__file__))  # 取得當前 app.py 的絕對目錄
# 確保上傳資料夾存在
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(os.path.join(basedir, 'json_data'), exist_ok=True)
os.makedirs(os.path.join(basedir, 'excel_result_Unclustered'), exist_ok=True)  # 新增未分群資料夾
os.makedirs(os.path.join(basedir, 'excel_result_Clustered'), exist_ok=True) # 新增分群資料夾

# ------------------------------------------------------------------------------

# 判斷是否允許的檔案格式
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ------------------------------------------------------------------------------

# 定義函數：處理特殊值（如 NaN、None 等）
def safe_value(val):
    if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
        return 0
    elif val is None:
        return None
    elif isinstance(val, str):
        return val
    else:
        return val

# ------------------------------------------------------------------------------

def append_progress(msg):
    global progress_log
    progress_log += msg + "\n"
    print(msg)  # 同時在控制台輸出
    
# ------------------------------------------------------------------------------




@app.route('/check-unclustered', methods=['GET'])
def check_unclustered_files():
    folder = 'excel_result_Unclustered'
    if not os.path.exists(folder):
        return jsonify({'exists': False}), 200
    files = [f for f in os.listdir(folder) if f.endswith('.xlsx')]
    return jsonify({'exists': len(files) > 0}), 200


@app.route('/clustered-files', methods=['GET'])
def list_clustered_files():
    clustered_folder = os.path.join('excel_result_Clustered', 'Details')

    if not os.path.exists(clustered_folder):
        return jsonify({'files': [], 'total': 0})

    # 支援分批參數
    offset = int(request.args.get('offset', 0))
    limit = int(request.args.get('limit', 100))

    pattern = re.compile(r"^Cluster_\[CI\].+_\[AI\].+\.xlsx$")  # 符合命名規則
    all_files = [
        f for f in os.listdir(clustered_folder)
        if f.endswith('.xlsx') and pattern.match(f)
    ]
    all_files.sort()  # 你也可以按日期或筆數排序

    total = len(all_files)  # 全部檔案總數

    # 分批取出本次要回傳的檔案
    page_files = all_files[offset:offset + limit]

    files_info = []
    for f in page_files:
        filepath = os.path.join(clustered_folder, f)
        try:
            df = pd.read_excel(filepath)
            row_count = len(df)
        except Exception as e:
            print(f"❌ 無法讀取 {f}：{e}")
            row_count = 0

        files_info.append({
            'name': f,
            'rows': row_count
        })

    return jsonify({'files': files_info, 'total': total})


@app.route('/download-clustered', methods=['GET'])
def download_clustered_file():
    filename = request.args.get('file')
    # 一定要加 Details 子資料夾
    path = os.path.join('excel_result_Clustered', 'Details', filename)
    if os.path.exists(path):
        return send_file(path, as_attachment=True)
    return jsonify({'error': '找不到檔案'}), 404


# 根據分數判斷風險等級（支援 KMeans 分群）
kmeans_thresholds = None  # 全域變數，存儲 KMeans 分群門檻


def get_risk_level(score):
    global kmeans_thresholds
    level = ''

    if kmeans_thresholds and len(kmeans_thresholds) == 4:
        thresholds = sorted(kmeans_thresholds)
        if score >= thresholds[3]:
            level = '高風險'
        elif score >= thresholds[2]:
            level = '中風險'
        elif score >= thresholds[1]:
            level = '低風險'
        else:
            level = '忽略'
        print(f"📊 KMeans：impactScore: {score} → 分級：{level}（使用動態門檻）")
    else:
        if score >= 18:
            level = '高風險'
        elif score >= 12:
            level = '中風險'
        elif score >= 6:
            level = '低風險'
        else:
            level = '忽略'
        print(f"📊 固定門檻：impactScore: {score} → 分級：{level}")

    return level

# 在分析完成後自動設定 kmeans_thresholds
# （請放在 KMeans 分群完成後）
def set_kmeans_thresholds_from_centroids(centroids):
    global kmeans_thresholds
    kmeans_thresholds = sorted(centroids)
    print(f"✅ 已設定 KMeans 分群門檻（sorted）：{kmeans_thresholds}")

# ------------------------------------------------------------------------------

def append_cluster_progress(msg, progress, total):
    with open("cluster_progress.json", "w", encoding="utf-8") as f:
        json.dump({
            "progress": progress,
            "total": total,
            "status": msg
        }, f, ensure_ascii=False)


def clean_filename(name):
    # 移除 Windows 禁用字元 < > : " / \ | ? *
    return re.sub(r'[<>:"/\\|?*]', '_', str(name)).strip()



def load_or_create_category_json(ci_name):
    safe_ci_name = clean_filename(ci_name)
    category_path = f"cluster_excels/{safe_ci_name}_categories.json"
    if os.path.exists(category_path):
        print(f"🔄 已讀取分類記憶 JSON：{category_path}")
        with open(category_path, "r", encoding="utf-8") as f:
            return json.load(f), category_path
    else:
        print(f"🆕 尚無分類記憶，將建立新檔：{category_path}")
        return [], category_path  # 空分類表

    
    

def classify_summary_with_ai(summary, existing_categories, config_item):
    # ✅ 格式化成逗號分隔字串給 prompt 用
    formatted_categories = ', '.join([str(c.get("category", "")) for c in existing_categories])

    print("-----------------------------------------------------")
    print("📋 開始進行分類")
    print(f"  - Summary：{summary[:80]}{'...' if len(summary) > 80 else ''}")
    print(f"  - 現有分類清單：{formatted_categories if formatted_categories else '(無)'}")
    
    prompt = f"""You are an AI assistant helping to classify IT incident summaries.

    The purpose of this classification is to produce more fine-grained and meaningful categories than the Configuration Item itself, so that IT incidents can be managed and analyzed more effectively.

    IMPORTANT:
    - The category **must be more specific than the Configuration Item**.
    - Do NOT use the Configuration Item name or any similar/derived name as the category.
    - For example, if the Configuration Item is "Email System", the category should be things like "login failure", "delivery delay", "mailbox full", "spam filtering", NOT just "Email System".

    Please read the following summary and determine the most appropriate category name from the given list.
    If none of the existing categories match well, create a new concise category that best represents the issue.

    Configuration Item:
    {config_item}

    Incident Summary:
    {summary}

    Existing Categories:
    {formatted_categories}

    Output Rule:
    - Return only the final category name.
    - Do NOT include any explanation, formatting, or punctuation.
    - The response must be exactly one line with the category name only.

    Your answer:"""


    try:
        payload = {
            "prompt": prompt
        }
        print("  - 已送出 API 請求，等待分類結果...")
        response = requests.post(POWERAUTOMATE_CLASSIFY_URL, json=payload, timeout=120)
        result = response.json()
        print("  - API 回傳內容：", result)
        category = result.get("category", "").strip().lower()
        print(f"  - 分類結果：{category if category else '（uncategorized）'}")
        print("-----------------------------------------------------")
        return category if category else "uncategorized"
    except Exception as e:
        print(f"❌ 分類失敗（summary: {summary[:80]}{'...' if len(summary) > 80 else ''}）：", e)
        print("-----------------------------------------------------")
        return "uncategorized"
# ------------------------------------------------------------------------------

@app.route('/cluster-excel', methods=['POST'])
def cluster_excel():
    unclustered_dir = 'excel_result_Unclustered'
    clustered_dir = 'excel_result_Clustered'
    os.makedirs(clustered_dir, exist_ok=True)  # ✅ 確保 Clustered 資料夾存在

    files = [f for f in os.listdir(unclustered_dir) if f.endswith('_Unclustered.xlsx')]
    
    total_files = len(files)

    append_cluster_progress("⏳ 開始分群...", 0, total_files)
    
    
    
    print("="*60)
    print(f"🔍 偵測到 {len(files)} 筆待分群檔案：{files if files else '（無）'}")
    print("="*60)

    for i, filename in enumerate(files, 1):
        
        # 🟡 一開始就寫「正在處理...」
        append_cluster_progress(
            f"正在處理第 {i}/{total_files} 個檔案：{filename}",
            i - 1,
            total_files
        )
        
        uid = filename.replace('_Unclustered.xlsx', '')
        excel_path = os.path.join(unclustered_dir, filename)
        print(f"\n🟦 [{i}/{len(files)}] 開始處理檔案：{excel_path}")

        df = pd.read_excel(excel_path)
        print(f"  📑 共 {len(df)} 筆資料待分類")

        # ✅ 一筆一筆分類
        for idx, row in df.iterrows():
            config_item = row.get("configurationItem", "Unknown")
            summary = row.get("aiSummary", "").strip()
            print(f"    ├─ [{idx+1}/{len(df)}] 分類 configurationItem：{config_item} | Summary 前 40 字：{summary[:40]}{'...' if len(summary)>40 else ''}")

            # 🔁 讀取分類記憶 JSON
            categories, cat_path = load_or_create_category_json(config_item)

            # ✨ 丟給 GPT 分類，這裡要多帶 config_item
            category = classify_summary_with_ai(summary, categories, config_item)

            # ✅ 如果分類還沒出現在記憶清單，才加入
            if category not in [c["category"].lower() for c in categories]:  # ✅ 小寫比較避免重複
                new_cat = {"category": category}
                categories.append(new_cat)
                with open(cat_path, "w", encoding="utf-8") as f:
                    json.dump(categories, f, indent=2, ensure_ascii=False)
                print(f"    │   🆕 新分類已加入 {cat_path}：{category}")
            else:
                print(f"    │   分類 {category} 已存在記憶清單")

            # ✅ 寫入這筆事件的 aiCategory
            df.at[idx, "aiCategory"] = category.lower()  # ✅ 保險轉小寫

        # ✅ 覆蓋回原 Excel（Unclustered）
        df.to_excel(excel_path, index=False)
        print(f"  📄 已覆蓋更新 Unclustered Excel：{excel_path}")

        # ✅ 繼續照原邏輯分群（此時每筆都有 aiCategory）
        results = df.to_dict(orient='records')
        print(f"  🔗 進行分群並匯出至 Clustered ...")
        cluster_excel_export(results)

        # ✅ 搬到 Clustered 資料夾
        clustered_path = os.path.join(clustered_dir, uid + '_Clustered.xlsx')
        shutil.move(excel_path, clustered_path)
        print(f"  📁 已移動檔案並改名：{clustered_path}")
                # ...分類、分群、summary...
        # ↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓
        append_cluster_progress(
            f"已完成第 {i}/{total_files} 個檔案：{filename}",
            i,
            total_files
        )
        
        
        
    # 所有檔案處理完成，最後寫一次進度
    append_cluster_progress(f"🎉 所有分群 Excel 檔案已完成搬移與分群，共處理 {total_files} 檔！", total_files, total_files)

    print("="*60)
    print(f"🎉 所有分群 Excel 檔案已完成搬移與分群，共處理 {len(files)} 檔！")
    print("="*60)

    return jsonify({'message': f'已成功處理 {len(files)} 筆 Excel 檔案並完成分群'}), 200

# ------------------------------------------------------------------------------

def cluster_excel_export(results, export_dir="excel_result_Clustered/Details"):
    def clean(text):
        return re.sub(r'[^\w\-_.]', '_', str(text).strip())[:30] or "Unknown"

    print("\n🟦 [cluster_excel_export] 開始分群與匯出 ...")
    cluster_data = defaultdict(list)
    for r in results:
        config_item = r.get("configurationItem") or "Unknown"
        ai_category = r.get("aiCategory") or "未分類"
        cluster_key = f"{config_item}_{ai_category}"
        r['cluster'] = cluster_key
        cluster_data[cluster_key].append(r)

    os.makedirs(export_dir, exist_ok=True)
    print(f"📁 匯出資料夾路徑：{export_dir}")
    print(f"🔢 共發現 {len(cluster_data)} 個分群")
    
    # ====== 加這段取得目前總檔案數，讓進度合理 ======
    total_clusters = len(cluster_data)
    # ==========================================
    

    for idx, (key, group) in enumerate(cluster_data.items(), 1):
        
        

        
        cluster_df = pd.DataFrame(group)
        try:
            config_item, ai_category = key.split('_', 1)
        except ValueError:
            config_item, ai_category = key, "未分類"

        filename = f"{export_dir}/Cluster_[CI]{clean(config_item)}_[AI]{clean(ai_category)}.xlsx"
        print(f"\n  [{idx}/{len(cluster_data)}] 分群 Key：{key}")
        print(f"    ├─ 匯出檔名：{filename}")
        print(f"    ├─ 本群共 {len(group)} 筆資料")

        if os.path.exists(filename):
            old_df = pd.read_excel(filename)
            print(f"    ├─ 檔案已存在，原有 {len(old_df)} 筆資料，將合併")
            cluster_df = pd.concat([old_df, cluster_df], ignore_index=True)
        else:
            print("    ├─ 檔案不存在，將新建檔案")

        cluster_df = cluster_df.sort_values(by="analysisTime", ascending=False)
        cluster_df.to_excel(filename, index=False)
        print(f"    ├─ 已輸出 Excel，合併後共 {len(cluster_df)} 筆")
        
        # ====== 新增：再存一份到 OneDrive 目的資料夾 (這個資料夾在MSTC ITG sharepoint 網站上)======
        custom_export_dir = r"C:\Users\tachang\Microsoft\MSTC ITG - Timmy\IncidentAnalysis_Clustered_File\Details"
        os.makedirs(custom_export_dir, exist_ok=True)
        custom_filename = os.path.join(custom_export_dir, os.path.basename(filename))
        cluster_df.to_excel(custom_filename, index=False)
        print(f"    ├─ 已額外備份到：{custom_filename}")
        # ==============================================
        
        
        
        print(f"    ├─ 產生群組摘要 ...")
        summarize_group_to_excel(config_item, ai_category, group)

        high_count = sum(1 for e in group if e.get('riskLevel') == '高風險')
        total = len(group)
        if total > 0 and (high_count / total) >= 0.5:
            print(f"    🚨 預警：Cluster {key} 有 {high_count}/{total} 筆高風險事件")
    print("\n✅ [cluster_excel_export] 所有分群 Excel 檔案已儲存！")
    
    
    
    
    
# ------------------------------------------------------------------------------
def run_gpt_summary(text, instruction):
    if not text.strip():
        print("⚠️ 無內容可摘要")
        return "（無內容）"

    prompt = f"""{instruction}

    Content:
    {text}

    Please summarize concisely in one paragraph:"""
    
    print("🟦 [run_gpt_summary] 開始呼叫 GPT 摘要")
    print(f"  - 指令 instruction：{instruction}")
    print(f"  - 摘要內容前 50 字：{text[:50]}{'...' if len(text)>50 else ''}")

    try:
        response = requests.post(POWERAUTOMATE_SUMMARY_URL, json={"prompt": prompt}, timeout=120)
        result = response.json()
        summary = result.get("summary", "").strip()
        print(f"  - GPT 摘要結果：{summary[:80]}{'...' if len(summary)>80 else ''}")
        print("🟦 [run_gpt_summary] 完成\n")
        return summary
    except Exception as e:
        print("❌ 摘要失敗：", e)
        return "（摘要失敗）"


def summarize_group_to_excel(config_item, ai_category, group, output_dir="excel_result_Clustered/Summaries"):
    def clean(text):
        return re.sub(r'[^\w\-_.]', '_', str(text).strip())[:30] or "Unknown"
    os.makedirs(output_dir, exist_ok=True)

    print("\n🟩 [summarize_group_to_excel] 開始群組摘要 ...")
    print(f"  - configItem：{config_item}")
    print(f"  - aiCategory：{ai_category}")
    print(f"  - 群組事件數量：{len(group)}")

    summaries = [str(r.get("aiSummary", "")).strip() for r in group if r.get("aiSummary")]
    solutions = [str(r.get("solution", "")).strip() for r in group if r.get("solution")]

    summary_text = "\n".join(summaries)
    solution_text = "\n".join(solutions)

    print(f"  - 問題 summaries 條數：{len(summaries)}")
    print(f"  - solution 條數：{len(solutions)}")
    print("  - 正在呼叫 GPT 產生『summary_summary』 ...")
    summary_summary = run_gpt_summary(summary_text, "Summarize the problem summaries in a concise form.")
    print("  - 正在呼叫 GPT 產生『solution_summary』 ...")
    solution_summary = run_gpt_summary(solution_text, "Summarize the solutions in a concise and helpful form.")

    # ✅ 儲存成 Excel
    df = pd.DataFrame([{
        "category": ai_category,
        "configItem": config_item,
        "summary_summary": summary_summary,
        "solution_summary": solution_summary
    }])
    # 確保輸出目錄存在
    filename = f"{output_dir}/Summary_[CI]{clean(config_item)}_[AI]{clean(ai_category)}.xlsx"

    df.to_excel(filename, index=False)
    print(f"📄 已輸出摘要檔案：{filename}")
    
    # ====== 新增：再存一份到 OneDrive 目的資料夾 (這個資料夾在MSTC ITG sharepoint 網站上)======
    custom_export_dir = r"C:\Users\tachang\Microsoft\MSTC ITG - Timmy\IncidentAnalysis_Clustered_File\Summaries"
    os.makedirs(custom_export_dir, exist_ok=True)
    custom_filename = os.path.join(custom_export_dir, os.path.basename(filename))
    df.to_excel(custom_filename, index=False)
    print(f"📄 已額外備份摘要到：{custom_filename}")
    # ==================================================    

    print("🟩 [summarize_group_to_excel] 完成\n")






@app.route('/download-summary', methods=['GET'])
def download_summary():
    filename = request.args.get('file')
    path = os.path.join('excel_result_Clustered', 'Summaries', filename)
    if os.path.exists(path):
        return send_file(path, as_attachment=True)
    return jsonify({'error': '找不到檔案'}), 404


@app.route('/summary-files', methods=['GET'])
def list_summary_files():
    # 確保 summary 資料夾存在
    summary_folder = os.path.join('excel_result_Clustered', 'Summaries')
    if not os.path.exists(summary_folder):
        return jsonify({'files': [], 'total': 0})

    # 取得分批參數，預設第一次 offset=0, limit=100
    offset = int(request.args.get('offset', 0))
    limit = int(request.args.get('limit', 100))

    # 取得全部符合條件的 .xlsx 檔案（建議按檔名排序，方便用戶找）
    all_files = [
        f for f in os.listdir(summary_folder)
        if f.endswith('.xlsx')
    ]
    all_files.sort()  # 你也可以改成用日期、大小排序

    total = len(all_files)  # 全部的 summary 檔案數
    page_files = all_files[offset:offset + limit]  # 這次要回傳的這一批

    files_info = []
    for f in page_files:
        filepath = os.path.join(summary_folder, f)
        try:
            df = pd.read_excel(filepath)
            row_count = len(df)
        except Exception as e:
            print(f"❌ 無法讀取 {f}：{e}")
            row_count = 0
        files_info.append({
            'name': f,
            'rows': row_count
        })

    return jsonify({'files': files_info, 'total': total})




# ------------------------------------------------------------------------------



# 用於同步 Flask 路由呼叫 async 分析邏輯
def analyze_excel(filepath, weights=None, resolution_priority=None, summary_priority=None):
    return asyncio.run(analyze_excel_async(filepath, weights, resolution_priority, summary_priority))



# 用於同步 Flask 路由呼叫 async 分析邏輯
async def analyze_excel_async(filepath, weights=None, resolution_priority=None, summary_priority=None):
    start_time = time.time()
    append_progress("🟩 開始分析 Excel 檔案...")
    default_weights = {
        'keyword': 5.0,
        'multi_user': 3.0,
        'escalation': 2.0,
        'config_item': 5.0,
        'role_component': 3.0,
        'time_cluster': 2.0
    }
    weights = {**default_weights, **(weights or {})}
    print(f"🟩 本次分析開始，將即時讀取三類語句 json 檔案...")
    # ⭐ 讀取語句和 embedding
    high_risk_examples, high_risk_embeddings = load_embeddings("high_risk")
    escalation_examples, escalation_embeddings = load_embeddings("escalate")
    multi_user_examples, multi_user_embeddings = load_embeddings("multi_user")
    append_progress("✅ 語句庫與嵌入向量載入完成")


    df = pd.read_excel(filepath)
    append_progress(f"✅ 成功載入檔案，共 {len(df)} 筆資料，開始欄位前處理")
    


    # ✅ 欄位順位 fallback 預設
    resolution_priority = resolution_priority or ['Description', 'Short description', 'Close notes']
    summary_priority = summary_priority or ['Short description', 'Description']

    def combine_fields_with_priority(row, field_order, limit):
        parts = []
        for f in field_order:
            if f in row and pd.notna(row[f]):
                value = str(row[f]).strip()
                parts.append(f"{f}: {value}")
        combined = "\n".join(parts)
        while len(combined) > limit and len(parts) > 1:
            removed = parts.pop()
            print(f"🔁 移除欄位：{removed[:20]}...")
            combined = "\n".join(parts)

        # 額外：印出實際使用的欄位名稱
        used_fields = field_order[:len(parts)]
        print(f"✅ 實際使用欄位：{used_fields}，合併長度：{len(combined)}")
        print("🔎 合併結果內容：\n", combined)   # 🟢 新增這行！
        return combined.strip()

    # ✅ 產生 resolution_input / summary_input 給 GPT 用
    df['resolution_input'] = df.apply(lambda row: combine_fields_with_priority(row, resolution_priority, 16000), axis=1)
    df['summary_input'] = df.apply(lambda row: combine_fields_with_priority(row, summary_priority, 16000), axis=1)
    append_progress("✅ Resolution / Summary 欄位內容合併完畢，準備進行逐列分析")



    component_counts = df['Role/Component'].value_counts()
    configuration_item_counts = df['Configuration item'].value_counts()
    configuration_item_max = configuration_item_counts.max()
    df['Opened'] = pd.to_datetime(df['Opened'], errors='coerce')
    analysis_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    append_progress("⏳ 正在分析所有資料，請稍候…")
    print(f"📊 角色/組件數量：{len(component_counts)}，最大值：{configuration_item_max}")
    # 非同步處理
    tasks = [
        analyze_row_async(
            row, idx, df, weights, component_counts, configuration_item_counts, configuration_item_max, analysis_time,
            high_risk_examples, high_risk_embeddings,
            escalation_examples, escalation_embeddings,
            multi_user_examples, multi_user_embeddings,
            row['resolution_input'], row['summary_input']  # ✅ 新增這兩欄
        )
        for idx, row in df.iterrows()
    ]
    # results_raw = await asyncio.gather(*tasks, return_exceptions=True)
    # results = [r for r in results_raw if r and not isinstance(r, Exception)]
    
    results = []
    total = len(tasks)
    finished = 0
    for coro in asyncio.as_completed(tasks):
        res = await coro
        finished += 1
        append_progress(f"🔄 分析進度：第 {finished} / {total} 筆")
        if res:
            results.append(res)

    # ✅ 防呆：沒有任何成功的結果就直接回傳避免崩潰
    if not results:
        print("⚠️ 所有資料都無法分析，請檢查欄位是否缺失")
        return {
            'data': [],
            'analysisTime': analysis_time
        }

    all_scores = [r['impactScore'] for r in results]
    score_range = max(all_scores) - min(all_scores)
    score_std = np.std(all_scores)

    print(f"📈 分群判斷指標：count={len(all_scores)}, range={score_range:.2f}, stddev={score_std:.2f}")

    if (
        len(all_scores) >= KMEANS_MIN_COUNT and
        score_range >= KMEANS_MIN_RANGE and
        score_std >= KMEANS_MIN_STDDEV
    ):
        kmeans = KMeans(n_clusters=4, random_state=42)
        labels = kmeans.fit_predict(np.array(all_scores).reshape(-1, 1))
        centroids = kmeans.cluster_centers_.flatten()
        set_kmeans_thresholds_from_centroids(centroids)
        print(f"📊 KMeans 分群標籤：{labels}")
        label_map = {}
        for i, idx in enumerate(np.argsort(centroids)[::-1]):
            label_map[idx] = ['高風險', '中風險', '低風險', '忽略'][i]
        for i, r in enumerate(results):
            r['riskLevel'] = label_map[labels[i]]
        print(f"📌 KMeans 分群中心：{sorted(centroids, reverse=True)}")
        
        append_progress(f"📈 KMeans 分群完成，中心值: {sorted(centroids, reverse=True)}")

    else:
        print("⚠️ 不啟用 KMeans，改用固定門檻分級")
        append_progress("⚠️ KMeans 未啟用，使用固定門檻進行分級")

        for r in results:
            r['riskLevel'] = get_risk_level(r['impactScore'])

    total_time = time.time() - start_time
    avg_time = total_time / len(results)
    append_progress(f"✅ 分析完成！總耗時 {total_time:.2f} 秒，平均每筆 {avg_time:.2f} 秒")
    print(f"\n🎯 所有分析總耗時：{total_time:.2f} 秒")
    print(f"📊 單筆平均耗時：{avg_time:.2f} 秒")
    print("\n✅ 所有資料分析完成！")
    return {
        'data': results,
        'analysisTime': analysis_time
    }
    
    
async def analyze_row_async(row, idx, df, weights, component_counts, configuration_item_counts, configuration_item_max, analysis_time,     
    high_risk_examples, high_risk_embeddings,
    escalation_examples, escalation_embeddings,
    multi_user_examples, multi_user_embeddings,
    resolution_text, summary_input):
    print(f"[分析 Row#{idx+1}] 本次用的高風險語句數：{len(high_risk_examples)}，倒數兩句：{high_risk_examples[-2:] if high_risk_examples else '空'}")
    print(f"[分析 Row#{idx+1}] 本次用的升級語句數：{len(escalation_examples)}，倒數兩句：{escalation_examples[-2:] if escalation_examples else '空'}")
    print(f"[分析 Row#{idx+1}] 本次用的影響多使用者語句數：{len(multi_user_examples)}，倒數兩句：{multi_user_examples[-2:] if multi_user_examples else '空'}")
    
    
    
    
    try:
        # 原始欄位保留
        description_text = row.get('Description', 'not filled')
        short_description_text = row.get('Short Description', 'not filled')
        close_note_text = row.get('Close notes', 'not filled')

        # 字串清理（保留變數命名）
        desc = str(description_text).strip()
        short_desc = str(short_description_text).strip()
        close_notes = str(close_note_text).strip()
        print("-------------------------------------------------------------------test----------------------------------------------------------------------------------------")
        print(desc, short_desc, close_notes)

        # 若全部內容皆為空，直接跳過此筆
        if not (desc or short_desc or close_notes):
            print(f"⚠️ 第 {idx+1} 筆內容全為空白，略過分析")
            return None

        # ✅ 改用合併後欄位（已由前段 fallback 處理）
        print(f"🧠 [Row#{idx+1}] 使用 resolution_text（長度：{len(resolution_text)}）")
        print(f"📌 Resolution 欄位原始合併內容：\n{resolution_text[:1000]}")
        print(f"🧠 [Row#{idx+1}] 使用 summary_input（長度：{len(summary_input)}）")
        print(f"📌 Summary 欄位原始合併內容：\n{summary_input[:1000]}")
        print("-------------------------------------------------------------------test----------------------------------------------------------------------------------------")
        print(desc, short_desc, close_notes)


        keyword_score = is_high_risk(summary_input, high_risk_examples, high_risk_embeddings)
        user_impact_score = is_multi_user(summary_input, multi_user_examples, multi_user_embeddings)
        escalation_score = is_escalated(resolution_text, escalation_examples, escalation_embeddings)

        config_raw = configuration_item_counts.get(row.get('Configuration item'), 0)
        configuration_item_freq = config_raw / configuration_item_max if configuration_item_max > 0 else 0

        role_comp = row.get('Role/Component', 'not filled')
        count = component_counts.get(role_comp, 0)
        role_component_freq = 3 if count >= 5 else 2 if count >= 3 else 1 if count == 2 else 0

        this_time = row.get('Opened', 'not filled')
        if pd.isnull(this_time):
            time_cluster_score = 1
        else:
            others = df[df['Role/Component'] == role_comp]
            close_events = others[(others['Opened'] >= this_time - pd.Timedelta(hours=24)) &
                                  (others['Opened'] <= this_time + pd.Timedelta(hours=24))]
            count_cluster = len(close_events)
            time_cluster_score = 3 if count_cluster >= 3 else 2 if count_cluster == 2 else 1

        severity_score = round(
            keyword_score * weights['keyword'] +
            user_impact_score * weights['multi_user'] +
            escalation_score * weights['escalation'], 2
        )
        frequency_score = round(
            configuration_item_freq * weights['config_item'] +
            role_component_freq * weights['role_component'] +
            time_cluster_score * weights['time_cluster'], 2
        )
        impact_score = round(math.sqrt(severity_score**2 + frequency_score**2), 2)
        risk_level = get_risk_level(impact_score)
        
        
        
        ai_suggestion, ai_summary = await analyze_with_ai_builder_then_fallback(
        resolution_text, summary_input, source_id=f"Row#{idx+1}"
        )

        return {
            'id': safe_value(row.get('Incident') or row.get('Number')),
            'configurationItem': safe_value(row.get('Configuration item')),
            'roleComponent': safe_value(row.get('Role/Component')),
            'subcategory': safe_value(row.get('Subcategory')),
            'aiSummary': safe_value(ai_summary),
            'originalShortDescription': safe_value(short_desc),
            'originalDescription': safe_value(desc),
            'severityScore': safe_value(severity_score),
            'frequencyScore': safe_value(frequency_score),
            'impactScore': safe_value(impact_score),
            'severityScoreNorm': round(severity_score / 10, 2),
            'frequencyScoreNorm': round(frequency_score / 20, 2),
            'impactScoreNorm': round(impact_score / 30, 2),
            'riskLevel': risk_level,
            'solution': safe_value(ai_suggestion or '無提供解法'),
            'location': safe_value(row.get('Location')),
            'opened': row.get('Opened'),  # ✅ 新增這行
            'analysisTime': analysis_time,
            'weights': {k: round(v / 10, 2) for k, v in weights.items()},
            'usedResolutionInput': safe_value(resolution_text),
            'usedSummaryInput': safe_value(summary_input),

        }

    except Exception as e:
        append_progress(f"❌ 分析第 {idx+1} 筆失敗：{str(e)[:60]}")
        print(f"❌ 分析第 {idx + 1} 筆失敗：", e)
        return None





@app.route('/cluster-progress')
def cluster_progress():
    try:
        with open("cluster_progress.json", encoding="utf-8") as f:
            return jsonify(json.load(f))
    except Exception:
        return jsonify({"progress": 0, "total": 1, "status": "尚未開始"})




SENTENCE_DIR = os.path.join("data", "sentences")
os.makedirs(SENTENCE_DIR, exist_ok=True)

def get_file_path(tag):
    return os.path.join(SENTENCE_DIR, f"{tag}.json")

@app.route("/get-sentence-db")
def get_sentence_db():
    result = []
    for tag in ["high_risk", "escalate", "multi_user"]:
        path = get_file_path(tag)
        if os.path.exists(path):
            with open(path, encoding='utf-8') as f:
                sentences = json.load(f)
                for entry in sentences:
                    result.append({"text": entry["text"], "tag": tag})
    return jsonify(result)

@app.route("/save-sentence-db", methods=["POST"])
def save_sentence():
    new_entry = request.get_json()
    tag = new_entry.get("tag")
    if tag not in ["high_risk", "escalate", "multi_user"]:
        return jsonify({"message": "invalid tag"}), 400

    path = get_file_path(tag)
    if os.path.exists(path):
        with open(path, encoding='utf-8') as f:
            data = json.load(f)
    else:
        data = []

    if any(d['text'] == new_entry['text'] for d in data):
        return jsonify({"message": "duplicate"}), 409

    data.append({"text": new_entry['text']})
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return get_sentence_db()

@app.route("/delete-sentence", methods=["POST"])
def delete_sentence():
    req = request.get_json()
    tag = req.get("tag")
    text = req.get("text")
    if tag not in ["high_risk", "escalate", "multi_user"] or not text:
        return jsonify({"message": "invalid input"}), 400

    path = get_file_path(tag)
    if not os.path.exists(path):
        return jsonify({"message": "not found"}), 404

    with open(path, encoding='utf-8') as f:
        data = json.load(f)

    new_data = [d for d in data if d['text'] != text]
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(new_data, f, ensure_ascii=False, indent=2)

    return get_sentence_db()

@app.route("/edit-sentence", methods=["POST"])
def edit_sentence():
    req = request.get_json()
    tag = req.get("tag")
    old_text = req.get("oldText")
    new_text = req.get("newText")

    if tag not in ["high_risk", "escalate", "multi_user"] or not old_text or not new_text:
        return jsonify({"message": "invalid input"}), 400

    path = get_file_path(tag)
    if not os.path.exists(path):
        return jsonify({"message": "not found"}), 404

    with open(path, encoding='utf-8') as f:
        data = json.load(f)

    updated = False
    for d in data:
        if d['text'] == old_text:
            d['text'] = new_text
            updated = True
            break

    if not updated:
        return jsonify({"message": "not found"}), 404

    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return get_sentence_db()








GPT_DATA_DIR = "gpt_data"
PROMPT_FILE = os.path.join(GPT_DATA_DIR, "gpt_prompts.json")
MAP_FILE = os.path.join(GPT_DATA_DIR, "gpt_prompt_map.json")

os.makedirs(GPT_DATA_DIR, exist_ok=True)

def read_json(path, default=None):
    if os.path.exists(path):
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    return default if default is not None else {}

def write_json(path, obj):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)

@app.route('/get-gpt-prompts')
def get_gpt_prompts():
    all_data = read_json(PROMPT_FILE, {})
    # 包成 {用途: {"prompts": [...]} }
    wrapped = {k: {"prompts": v if isinstance(v, list) else [v]} for k, v in all_data.items()}
    return jsonify(wrapped)

@app.route('/save-gpt-prompt', methods=['POST'])
def save_gpt_prompt():
    """
    新增一筆 prompt 到某個分類。body: { "task": "solution", "prompt": "xxx" }
    """
    data = request.get_json()
    task = data.get('task')
    prompt = data.get('prompt', '').strip()

    if not task or not prompt:
        return jsonify(success=False, message='❌ 缺少 task 或 prompt'), 400

    all_prompts = read_json(PROMPT_FILE, {})
    prompt_list = all_prompts.get(task, [])
    if not isinstance(prompt_list, list):
        prompt_list = [prompt_list]

    if prompt not in prompt_list:
        prompt_list.append(prompt)
    else:
        return jsonify(success=False, message='⚠️ 該 prompt 已存在'), 409

    all_prompts[task] = prompt_list
    write_json(PROMPT_FILE, all_prompts)

    return jsonify(success=True, allPrompts=all_prompts)

@app.route('/delete-gpt-prompt', methods=['POST'])
def delete_gpt_prompt():
    """
    支援刪除分類或分類下的單一句子。
    body: { "task": "solution", "prompt": "xxx" } 或只給 task 代表整類刪除。
    """
    data = request.get_json()
    task = data.get('task')
    prompt = data.get('prompt', '').strip()

    all_prompts = read_json(PROMPT_FILE, {})

    if task not in all_prompts:
        return jsonify(success=False, message=f'找不到用途 {task}'), 404

    if prompt:
        prompt_list = all_prompts[task]
        if prompt in prompt_list:
            prompt_list.remove(prompt)
            if prompt_list:
                all_prompts[task] = prompt_list
            else:
                del all_prompts[task]
        else:
            return jsonify(success=False, message='找不到該 prompt'), 404
    else:
        del all_prompts[task]  # 刪整類

    write_json(PROMPT_FILE, all_prompts)

    # 刪掉 mapping 中的對應
    mapping = read_json(MAP_FILE, {})
    if task in mapping:
        del mapping[task]
        write_json(MAP_FILE, mapping)

    return jsonify(success=True, allPrompts=all_prompts)

@app.route('/get-gpt-prompt-map')
def get_gpt_prompt_map():
    return jsonify(read_json(MAP_FILE, {}))

@app.route('/save-gpt-prompt-map', methods=['POST'])
def save_gpt_prompt_map():
    data = request.get_json()

    solution_prompt = data.get("solution")
    summary_prompt = data.get("ai_summary")
    models = data.get("models", {})

    new_mapping = {
        "solution": {
            "prompt": solution_prompt,
            "model": models.get("solution", "")
        },
        "ai_summary": {
            "prompt": summary_prompt,
            "model": models.get("ai_summary", "")
        }
    }

    write_json(MAP_FILE, new_mapping)
    return jsonify(success=True, mapping=new_mapping)

def get_prompt_for_use(use_type):
    mapping = read_json(MAP_FILE, {})
    all_prompts = read_json(PROMPT_FILE, {})
    mapped_key = mapping.get(use_type, use_type)
    prompt_list = all_prompts.get(mapped_key, [])
    if isinstance(prompt_list, list):
        return {'prompt': prompt_list[0] if prompt_list else '', 'model': mapping.get(mapped_key, {}).get("model", "")}
    return {'prompt': prompt_list, 'model': mapping.get(mapped_key, {}).get("model", "")}

# 定義首頁路由
@app.route('/')
def index():
    return render_template('FrontEnd.html')  # 渲染首頁模板

# 定義結果頁面路由
@app.route('/result')
def result_page():
    return render_template('result.html')  # 渲染結果頁面

# 定義歷史紀錄頁面路由
@app.route('/history')
def history_page():
    return render_template('history.html')  # 渲染歷史紀錄頁面


@app.route('/generate_cluster')
def generate_cluster_page():
    return render_template('generate_cluster.html')  # 渲染生成分群頁面

@app.route("/manual_input")
def manual_input_page():
    return render_template("manual_input.html")

@app.route("/gpt_prompt")
def gpt_prompt_page():
    return render_template("gpt_prompt.html")

@app.route("/chat_ui")
def helpdesk_ui():
    return render_template("chat.html")



# ------------------------------------------------------------------------------

@app.route("/chat", methods=["POST"])
def chat_with_model():
    data = request.get_json()
    message = data.get("message", "")
    model = data.get("model", "mistral") # model 預設是 mistral
    history = data.get("history", [])
    chat_id = data.get("chatId", "")  # ✅ 前端傳入的唯一 ID

    if not chat_id:
        return jsonify({"error": "Missing chatId"}), 400

    # 建立資料夾與檔案路徑
    os.makedirs("chat_history", exist_ok=True)
    file_path = os.path.join("chat_history", f"{chat_id}.json")

    try:
        # ✅ 呼叫 GPT 模型處理（你的核心邏輯）
        reply = asyncio.run(run_offline_gpt(message, model=model, history=history, chat_id=chat_id))
        
        print("🧾 reply 類型：", type(reply))
        print("🧾 reply 內容預覽：", str(reply)[:1000])

        # ✅ 安全處理：保證 reply 一定是 str，避免 list/dict 錯誤
        if not isinstance(reply, str):
            print("⚠️ reply 不是字串，自動轉換為安全格式")
            try:
                reply = json.dumps(reply, ensure_ascii=False)
            except Exception as e:
                reply = f"⚠️ 回傳格式錯誤：{e}"

        # ✅ 判斷是新話題還是繼續聊
        if not os.path.exists(file_path):
            print("🆕 新的對話紀錄，建立新檔案")
            # 🆕 首次建立新檔案
            chat_record = {
                "id": chat_id,
                "title": chat_id,     # 固定為檔名
                "edit_title": "",     # 預設空
                "model": model,
                "timestamp": datetime.now().isoformat(),
                "history": history + [
                    {"role": "user", "content": message},
                    {"role": "assistant", "content": reply}
                ]
            }
        else:
            # 🔁 載入原檔案並追加
            with open(file_path, "r", encoding="utf-8") as f: # 打開既有的對話紀錄檔案（JSON 格式）
                chat_record = json.load(f) 
                
            # 🛡️ 防呆：如果 history 是空的或格式錯誤，強制初始化為 list
            if "history" not in chat_record or not isinstance(chat_record["history"], list):
                print("⚠️ history 格式錯誤，自動初始化為空 list")
                chat_record["history"] = []

            chat_record["history"].append({"role": "user", "content": message}) # 追加使用者的訊息
            chat_record["history"].append({"role": "assistant", "content": reply}) # 追加助手的回覆
            print("📝 已載入既有對話紀錄，並追加新的訊息。")

        # ✅ 寫回檔案
        with open(file_path, "w", encoding="utf-8") as f: # 打開檔案準備寫入,如果檔案已存在，會覆蓋原內容。
            # 將對話紀錄寫入 JSON 檔案
            json.dump(chat_record, f, ensure_ascii=False, indent=2)

        print(f"✅ 對話紀錄已儲存到 {file_path}")
        return jsonify({"reply": reply}) # 回傳助手的回覆用json形式

    except Exception as e:
        print(f"❌ 發生錯誤test: {str(e)}")  # ✅ 加上錯誤內容
        return jsonify({"error": str(e)}), 500 # 如果發生錯誤，回傳錯誤訊息



@app.route("/rename-chat", methods=["POST"])
def rename_chat():
    data = request.get_json()
    chat_id = data.get("chatId")
    new_title = data.get("newTitle")

    if not chat_id or new_title is None:
        return jsonify({"error": "缺少參數"}), 400

    file_path = Path(f"chat_history/{chat_id}.json")
    if not file_path.exists():
        return jsonify({"error": "找不到檔案"}), 404

    try:
        with open(file_path, encoding="utf-8") as f:
            record = json.load(f)
        record["edit_title"] = new_title  # ✅ 寫入新標題
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(record, f, ensure_ascii=False, indent=2)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    

@app.route("/delete-chat/<chat_id>", methods=["DELETE"])
def delete_chat(chat_id):
    file_path = Path(f"chat_history/{chat_id}.json")
    if file_path.exists():
        try:
            file_path.unlink()  # 刪除檔案
            return jsonify({"success": True})
        except Exception as e:
            return jsonify({"error": f"無法刪除：{e}"}), 500
    return jsonify({"error": "檔案不存在"}), 404



@app.route("/chat-history-list")
def get_chat_history_list():
    records = []
    path = Path("chat_history")
    if not path.exists():
        return jsonify([])

    for file in path.glob("*.json"):
        with open(file, encoding="utf-8") as f:
            try:
                obj = json.load(f)
                records.append({
                    "id": obj.get("id"),
                    "title": obj.get("edit_title") or obj.get("title"),
                    "timestamp": obj.get("timestamp"),
                    "model": obj.get("model")
                })
            except:
                continue
    return jsonify(sorted(records, key=lambda x: x["timestamp"], reverse=True))

@app.route("/chat-history/<id>")
def get_chat_history_by_id(id):
    file_path = Path(f"chat_history/{id}.json")
    if not file_path.exists():
        return jsonify({"error": "not found"}), 404
    with open(file_path, encoding="utf-8") as f:
        return jsonify(json.load(f))




@app.route('/preview-excel', methods=['POST'])
def preview_excel():
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '未提供檔案'})

    try:
        df = pd.read_excel(file)
        columns = df.columns.tolist()
        rows = df.head(50).fillna('').astype(str).to_dict(orient='records')  # 預覽前 50 筆資料
        return jsonify({'columns': columns, 'rows': rows})
    except Exception as e:
        return jsonify({'error': str(e)})
    
    
@app.route('/get-progress')
def get_progress():
    global progress_log
    return jsonify({"progress": progress_log})






@app.route('/ping')
def ping():
    return "pong", 200


@app.route('/upload', methods=['POST'])
def upload_file():
    global progress_log
    progress_log = ""     # 每次新上傳都先清空進度
    append_progress("✅ 開始分析 ...")
    print("📥 收到上傳請求")

    if 'file' not in request.files:
        print("❌ 沒有 file 欄位")
        return jsonify({'error': '沒有找到檔案欄位'}), 400

    file = request.files['file']
    if file.filename == '':
        print("⚠️ 檔案名稱為空")
        return jsonify({'error': '未選擇檔案'}), 400

    if not allowed_file(file.filename):
        print("⚠️ 檔案類型不符")
        return jsonify({'error': '請上傳 .xlsx 檔案'}), 400

    # 接收權重設定
    weights = None
    weights_raw = request.form.get('weights')
    if weights_raw:
        try:
            weights = json.loads(weights_raw)
            print("📥 收到權重設定：", weights)
        except Exception as e:
            print(f"⚠️ 權重解析失敗：{e}")
            return jsonify({'error': '權重解析失敗'}), 400
    else:
        print("ℹ️ 未提供自訂權重，使用預設值分析")

    # 解析 resolution/summary 欄位順位
    try:
        resolution_priority = json.loads(request.form.get('resolution_priority', '[]'))
        summary_priority = json.loads(request.form.get('summary_priority', '[]'))
        print("📌 Resolution 順位欄位：", resolution_priority)
        print("📌 Summary 順位欄位：", summary_priority)
    except Exception as e:
        print(f"⚠️ 欄位順位解析失敗：{e}")
        return jsonify({'error': '欄位順位解析失敗'}), 400

    # 產生時間戳記與檔名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    uid = f"result_{timestamp}"
    original_filename = f"original_{timestamp}.xlsx"
    original_path = os.path.join('uploads', original_filename)

    try:
        file.save(original_path)
        print(f"📁 原始檔已儲存：{original_path}")
    except Exception as e:
        return jsonify({'error': f'儲存原始檔失敗：{str(e)}'}), 500

    try:
        # 呼叫分析主邏輯
        analysis_result = analyze_excel(
            original_path,
            weights=weights,
            resolution_priority=resolution_priority,
            summary_priority=summary_priority
        )
        results = analysis_result['data']
        # ⬇️⬇️⬇️ 就加在這！
        analysis_result['file'] = file.filename

        save_analysis_files(analysis_result, uid)
        print(f"✅ 分析完成，共 {len(results)} 筆")




        # 自動觸發建庫腳本
        print("🚀 自動執行 build_kb.py 建立知識庫")
        # 呼叫本地的 Python 執行 build_kb.py（保證和 Flask 用同一個解譯器）
        script_path = os.path.join(os.path.dirname(__file__), "build_kb.py")
        print("🚀 嘗試用 sys.executable 執行：", script_path)
        subprocess.Popen([sys.executable, script_path])

        # 這裡直接讀最新的分析檔
        json_path = os.path.join('json_data', f"{uid}.json")
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                loaded = json.load(f)
                results = loaded.get('data', [])
        else:
            results = []
        print(f"📁 最新分析檔已讀取：{json_path}")
        
        return jsonify({
            'data': results[:100],   # 只回傳前 100 筆
            'uid': uid,
            'weights': weights,
            'jsonFilename': f"{uid}.json"  # 新增這一行
        }), 200
    except Exception as e:
        print(f"❌ 分析時發生錯誤：{e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    


def make_json_serializable(obj):
    if isinstance(obj, pd.Timestamp):
        return obj.isoformat()
    if isinstance(obj, dict):
        return {k: make_json_serializable(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [make_json_serializable(v) for v in obj]
    return obj




import os
import pandas as pd
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook



def append_df_to_excel(df, excel_path, sheet_name="Sheet1"):
    if not os.path.exists(excel_path):
        print(f"📁 檔案不存在，將建立新檔案：{excel_path}")
        # 建立新檔 + 表格格式
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        wb = load_workbook(excel_path)
        ws = wb[sheet_name]
        table = Table(displayName="IncidentTable",
                      ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(table)
        wb.save(excel_path)
        print("✅ 新檔案已建立並表格化")
        return

    # 檔案存在：先讀現有資料列數與 workbook
    reader = pd.read_excel(excel_path, sheet_name=sheet_name, engine="openpyxl")
    book = load_workbook(excel_path)

    # 若 sheet 不存在，手動建立表頭+
    if sheet_name not in book.sheetnames:
        ws = book.create_sheet(title=sheet_name)
        for idx, col in enumerate(df.columns, start=1):
            ws.cell(row=1, column=idx, value=col)
        table = Table(displayName="IncidentTable",
                      ref=f"A1:{get_column_letter(len(df.columns))}1")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(table)
        book.save(excel_path)
        print("✅ 已建立表頭與表格樣式")

    startrow = len(reader)
    print(f"📌 目標檔案已有 {startrow} 列，從第 {startrow + 1} 列開始追加")

    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:

        # ✅ 使用 update() 而非重新賦值
        writer.sheets.update({ws.title: ws for ws in book.worksheets})
        df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False)
    print("✅ 成功 append 分析結果到目標 Excel！")

    # 📏 自動延展表格範圍（加強穩定性）
    try:
        wb2 = load_workbook(excel_path)
        ws2 = wb2[sheet_name]

        if ws2.tables:
            table = list(ws2.tables.values())[0]
            new_ref = f"A1:{get_column_letter(ws2.max_column)}{ws2.max_row}"
            print(f"📐 延展表格範圍：{table.ref} ➜ {new_ref}")
            table.ref = new_ref
            wb2.save(excel_path)
            print("📏 表格範圍已更新")
            
        else:
            print("⚠️ 找不到任何表格，將補建表格")
            new_ref = f"A1:{get_column_letter(ws2.max_column)}{ws2.max_row}"
            new_table = Table(displayName="IncidentTable", ref=new_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            new_table.tableStyleInfo = style
            ws2.add_table(new_table)
            wb2.save(excel_path)
            print(f"🆕 已補建表格並設定範圍：{new_ref}")
            
            
        # ✅ 🔁 再次讀取檔案，這次真正設定欄寬（確保先 append、再調整）
        wb3 = load_workbook(excel_path)
        ws3 = wb3[sheet_name]

        # 🧩 自動調整每一欄欄寬
        print("📐 正在自動調整欄寬...")
        for i in range(1, ws3.max_column + 1):
            max_width = 1
            for j in range(1, ws3.max_row + 1):
                cell_value = ws3.cell(row=j, column=i).value
                if isinstance(cell_value, (int, float)):
                    width = len(str(format(cell_value, ',')))
                elif cell_value is None:
                    width = 0
                else:
                    width = len(str(cell_value).encode('gbk'))  # 中文 1 字 = 2 bytes（gbk 編碼）
                max_width = max(max_width, width)
            col_letter = get_column_letter(i)
            ws3.column_dimensions[col_letter].width = min(max_width, 20) + 2  # 最多 20，加 2 顯得寬鬆

        wb3.save(excel_path)
        print("📐 表格與欄位寬度已更新")
        

    except Exception as e:
        print(f"⚠️ 延展表格範圍失敗：{e}")



def deduplicate_by_id_and_time(df: pd.DataFrame) -> pd.DataFrame:
    if "id" not in df.columns or "analysisTime" not in df.columns:
        print("⚠️ 缺少 id 或 analysisTime 欄位，無法去重")
        return df

    df["id"] = df["id"].astype(str).str.strip()
    df["analysisTime_parsed"] = pd.to_datetime(df["analysisTime"], errors="coerce", utc=True)

    before = len(df)
    df = df.dropna(subset=["id", "analysisTime_parsed"])
    df = df.sort_values("analysisTime_parsed").drop_duplicates(subset="id", keep="last")
    after = len(df)

    print(f"🧹 去重完成：原始 {before} 筆 ➜ 去重後 {after} 筆")
    return df.drop(columns=["analysisTime_parsed"])

# ✅ 欄位名稱對照：名稱 → 要送出的名稱
FIELD_MAPPING = {
    "id": "id",
    "configurationItem": "configurationItem",
    "roleComponent": "roleComponent",
    "subcategory": "subcategory",
    "aiSummary": "problem",  # ← 改這行
    "solution": "solution",
    "severityScoreNorm": "severityScoreNorm",   # ✅ 保留原名
    "frequencyScoreNorm": "frequencyScoreNorm",
    "impactScoreNorm": "impactScoreNorm",
    "riskLevel": "riskLevel",
    "location": "location",
    "opened": "opened",
    "analysisTime": "analysisTime"  # ✅ 別忘了加上時間    
}

def apply_excel_formatting(path, sheet_name="Sheet1"):
    wb = load_workbook(path)
    ws = wb[sheet_name]
    new_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # 📊 建立表格樣式
    table = Table(displayName="IncidentTable", ref=new_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # 📐 自動調整欄寬
    for i in range(1, ws.max_column + 1):
        max_width = 1
        for j in range(1, ws.max_row + 1):
            val = ws.cell(row=j, column=i).value
            if isinstance(val, (int, float)):
                width = len(str(format(val, ',')))
            else:
                width = len(str(val or "").encode('gbk'))
            max_width = max(max_width, width)
        ws.column_dimensions[get_column_letter(i)].width = min(max_width, 20) + 2

    wb.save(path)
    print("🎨 表格樣式與欄寬已自動調整")



def save_analysis_files(result, uid):
    os.makedirs('json_data', exist_ok=True)
    os.makedirs('excel_result_Unclustered', exist_ok=True)  # ✅ 使用新的資料夾

    # 儲存 JSON
    json_path = os.path.join(basedir, 'json_data', f"{uid}.json")
    print(f"📝 預計儲存 JSON：{json_path}")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(make_json_serializable(result), f, ensure_ascii=False, indent=2)
    print("✅ JSON 檔案已寫入成功")

    # 儲存分析報表 Excel（只儲存 result['data']）
    df = pd.DataFrame(result['data'])
    
    print(f"匯入Excel 檔案：{df}")

    # ✅ 僅保留並排序目標欄位（原始欄位名稱）
    columns_to_keep = list(FIELD_MAPPING.keys())
    for col in columns_to_keep:
        if col not in df.columns:
            df[col] = ""
    df = df[columns_to_keep]

    # ✅ 去重處理
    df = deduplicate_by_id_and_time(df)

    # ✅ 欄位轉換（左邊欄位 ➜ Sync 欄位名稱）
# ✅ 欄位轉換
    df = df.rename(columns=FIELD_MAPPING)

    # ✅ 按照你想要的順序保留欄位（右邊的欄位名稱）
    sync_columns_ordered = [FIELD_MAPPING[k] for k in FIELD_MAPPING if FIELD_MAPPING[k] in df.columns]
    df = df[sync_columns_ordered]

    
    # ✅ 儲存到 Unclustered 資料夾並加上 Unclustered 後綴
    excel_filename = f"{uid}_Unclustered.xlsx"
    excel_path = os.path.join(basedir, 'excel_result_Unclustered', excel_filename)    
    df.to_excel(excel_path, index=False)

    # 確認 JSON 檔案是否寫入成功
    if os.path.exists(json_path):
        print("✅ JSON 檔案已成功儲存")
    else:
        print("❌ JSON 檔案儲存失敗！")

    print(f"✅ 分析報表已儲存：{excel_path}")
    print("📁 JSON 絕對路徑：", os.path.abspath(json_path))
    print("📁 Excel 絕對路徑：", os.path.abspath(excel_path))
    timestamp = uid.replace("result_", "")
    original_excel_path = os.path.abspath(os.path.join(basedir, 'uploads', f"original_{timestamp}.xlsx"))

    # ✅ 附加到同步的 Excel 檔案（不影響原有儲存流程）
    try:
        sync_target = r"C:\Users\tachang\Microsoft\MSTC ITG - Timmy\IncidentAnalysisDB.xlsx"

        # ✅ 若資料夾不存在就先建立
        folder_path = os.path.dirname(sync_target)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
            print(f"📁 已自動建立資料夾：{folder_path}")

        if os.path.exists(sync_target):
            print("📖 讀取歷史 Excel 中資料進行合併與去重")
            old_df = pd.read_excel(sync_target, sheet_name="Sheet1", engine="openpyxl")
            print(f"📄 歷史資料筆數：{len(old_df)}")
            combined_df = pd.concat([old_df, df], ignore_index=True)
            combined_df = deduplicate_by_id_and_time(combined_df)
        else:
            print("🆕 無歷史檔案，直接使用本次資料")
            combined_df = df

        # ✅ 整體覆蓋寫入，不再使用 append
        with pd.ExcelWriter(sync_target, engine="openpyxl", mode="w") as writer:
            combined_df.to_excel(writer, sheet_name="Sheet1", index=False)

        # ✅ 資料寫入完成後，再進行格式套用
        apply_excel_formatting(sync_target)
        print(f"📊 準備覆蓋寫入 {sync_target}，共 {len(combined_df)} 筆資料")
        print(f"✅ 成功覆蓋寫入 Excel（共 {len(combined_df)} 筆）：{sync_target}")

    except PermissionError:
        print("⚠️ 無法寫入，請確認 Excel 檔案是否已關閉！")
    except Exception as e:
        print(f"⚠️ 無法寫入 Excel：{e}")
        traceback.print_exc()

        
        
        
@app.route("/compare-file", methods=["POST"])
def compare_file():
    print("📥 收到檔案比對請求") 
    uploaded_file = request.files.get("file")
    if not uploaded_file:
        print("❌ 沒有收到上傳的檔案")
        return jsonify({"error": "No file uploaded"}), 400
 
    try:
        # ✅ 建立暫存資料夾
        temp_dir = "tmp_upload"
        os.makedirs(temp_dir, exist_ok=True)
        print(f"📁 確保暫存資料夾存在：{temp_dir}")
 
        # ✅ 儲存上傳檔案
        temp_path = os.path.join(temp_dir, uploaded_file.filename)
        uploaded_file.save(temp_path)
        print(f"📄 已儲存上傳檔案至暫存：{temp_path}")
 
        # ✅ 讀取上傳的檔案
        df_new = pd.read_excel(temp_path)
        print("📊 成功讀取上傳檔案為 DataFrame")
 
        # ✅ 比對 uploads 中的檔案
        uploads_dir = "uploads"
        print(f"🔍 開始比對 uploads 資料夾內檔案，共 {len(os.listdir(uploads_dir))} 個")
 
        for fname in os.listdir(uploads_dir):
            fpath = os.path.join(uploads_dir, fname)
            try:
                df_existing = pd.read_excel(fpath)
                print(f"📎 正在比對：{fname}")
                if df_new.equals(df_existing):
                    print(f"✅ 發現重複檔案：{fname}")
                    os.remove(temp_path)
                    print(f"🧹 已刪除暫存檔案：{temp_path}")
                    return jsonify({"duplicate": True})
            except Exception as ex:
                print(f"⚠️ 無法讀取檔案 {fname}：{ex}")
                continue
 
        os.remove(temp_path)
        print(f"🧹 比對完成，未發現重複。已刪除暫存檔案：{temp_path}")
        return jsonify({"duplicate": False})
 
    except Exception as e:
        if 'temp_path' in locals() and os.path.exists(temp_path):
            os.remove(temp_path)
            print(f"❗發生錯誤，刪除暫存檔案：{temp_path}")
        print(f"❌ 發生錯誤：{e}")
        return jsonify({"error": str(e)}), 500
 



@app.route('/kb-status')
def kb_status():
    lock_exists = os.path.exists("kb_building.lock")
    print(f"[DEBUG] lock file exists? {lock_exists}")
    return jsonify({"building": lock_exists})


from datetime import datetime, timedelta


@app.route('/get-results')
def get_results():
    folder = 'json_data'
    results = []
    first_weights = {}
        # ✅ 新增：取得前端傳來的參數（預設 start=0, limit=20）
    start = int(request.args.get('start', 0))
    limit = int(request.args.get('limit', 20))
    filter_days = request.args.get('days', 'all')

    if not os.path.exists(folder):
        return jsonify({'error': f'資料夾不存在：{folder}'}), 404

    # 🔄 讀取所有檔案，找出最新的那份分析檔
    sorted_files = sorted(
        [f for f in os.listdir(folder) if f.endswith('.json')],
        reverse=True  # 最後面最新
    )

    for filename in sorted_files:
        filepath = os.path.join(folder, filename)
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                content = json.load(f)
                if isinstance(content, dict) and 'data' in content:
                    results.extend(content['data'])
                    if not first_weights and 'weights' in content:
                        first_weights = content['weights']
                elif isinstance(content, list):
                    results.extend(content)
        except Exception as e:
            print(f"❌ 錯誤讀取 {filename}：{e}")
            
            
            
    # ✅ 新增：依據篩選天數過濾資料
    if filter_days != 'all':
        try:
            days = int(filter_days)
            now = datetime.now()
            if days == 0:
                # 只顯示今天，threshold 是今天的日期
                threshold_date = now.date()
            else:
                threshold_date = (now - timedelta(days=days)).date()
            results = [
                r for r in results
                if 'analysisTime' in r
                and isinstance(r['analysisTime'], str)
                and datetime.fromisoformat(r['analysisTime']).date() >= threshold_date
            ]
        except:
            pass


    # ✅ 新增：分批截取要傳回的資料（從第 start 筆取 limit 筆）
    sliced = results[start:start + limit]

    return jsonify({
        'data': sliced,  # ✅ 傳回當前批次的資料
        'total': len(results),     # ✅ 傳回總筆數給前端判斷是否載完
        'weights': first_weights  # ✅ 確保傳出這個欄位
    })



HISTORY_FOLDER = 'json_data'  # 你的 JSON 檔案存放資料夾


@app.route('/history-list', methods=['GET'])
def get_history_list():
    # 取得分頁參數，預設 page=1, pageSize=10
    page = int(request.args.get('page', 1))
    pageSize = int(request.args.get('pageSize', 10))
    records = []

    if not os.path.exists(HISTORY_FOLDER):
        # 多回傳 total 讓前端知道是 0 筆
        return jsonify({"records": [], "total": 0})

    # 先把所有檔案依時間排序（新到舊）
    all_files = sorted(
        [f for f in os.listdir(HISTORY_FOLDER) if f.endswith('.json')],
        reverse=True
    )
    total = len(all_files)  # 總共幾筆檔案

    # 算出要拿第幾筆到第幾筆
    start_idx = (page - 1) * pageSize
    end_idx = start_idx + pageSize
    page_files = all_files[start_idx:end_idx]

    for fname in page_files:
        file_path = os.path.join(HISTORY_FOLDER, fname)
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                record = json.load(f)
                if "data" in record and isinstance(record["data"], list):
                    risk_levels = [row.get("riskLevel", "未知") for row in record["data"]]
                    level_count = Counter(risk_levels)
                    risk_str = "｜".join([
                        f"{label}:{level_count[label]}"
                        for label in ["高風險", "中風險", "低風險", "忽略"]
                        if level_count.get(label, 0) > 0
                    ])
                else:
                    risk_str = "—"
                base_summary = record.get("summary", "")
                final_summary = risk_str if not base_summary else risk_str + "｜" + base_summary

                records.append({
                    "uid": record.get("uid", fname.replace('.json','')),
                    "file": record.get("file", fname),
                    "summary": final_summary,
                    "time": record.get("analysisTime", record.get("time", "未知時間"))
                })
        except Exception as e:
            print(f"❌ 解析歷史檔錯誤：{fname}", e)

    # 回傳本頁資料和總數
    return jsonify({
        "records": records,    # 當前這一頁的資料（list）
        "total": total         # 所有檔案總數
    })
    
    


@app.route('/clear-history', methods=['POST'])
def clear_history():
    if not os.path.exists(HISTORY_FOLDER):
        return jsonify({"success": True})

    deleted = 0
    for fname in os.listdir(HISTORY_FOLDER):
        if fname.endswith('.json'):
            try:
                os.remove(os.path.join(HISTORY_FOLDER, fname))
                deleted += 1
            except Exception as e:
                print(f"❌ 無法刪除 {fname}", e)
    return jsonify({"success": True, "deleted": deleted})










# ✅ JSON 預覽路由：提供 `/get-json?file=xxxx.json`
@app.route('/get-json', methods=['GET'])
def get_json_file():
    filename = request.args.get('file')  # e.g., result_20250423_152301.json
    if not filename:
        return jsonify({'error': '缺少 file 參數'}), 400

    json_path = os.path.join('json_data', filename)
    if os.path.exists(json_path):
        return send_file(json_path, as_attachment=False)
    else:
        return jsonify({'error': '找不到對應的 JSON 檔案'}), 404


@app.route('/download-excel', methods=['GET'])
def download_excel_file():
    uid = request.args.get('uid')  # e.g., result_20250508_203611
    if not uid:
        return jsonify({'error': '缺少 uid 參數'}), 400

    # 先檢查 Clustered
    clustered_path = os.path.join('excel_result_Clustered', f"{uid}_Clustered.xlsx")
    if os.path.exists(clustered_path):
        return send_file(clustered_path, as_attachment=True)

    # 再檢查 Unclustered
    unclustered_path = os.path.join('excel_result_Unclustered', f"{uid}_Unclustered.xlsx")
    if os.path.exists(unclustered_path):
        return send_file(unclustered_path, as_attachment=True)

    return jsonify({'error': f'找不到 {uid} 對應的 Excel 檔案'}), 404


@app.route('/download-original', methods=['GET'])
def download_original_excel():
    uid = request.args.get('uid')  # uid = result_20250423_152301
    if not uid:
        return jsonify({'error': '缺少 uid 參數'}), 400

    # 取出對應的時間戳
    timestamp = uid.replace('result_', '')
    original_filename = f'original_{timestamp}.xlsx'
    original_path = os.path.join('uploads', original_filename)

    if os.path.exists(original_path):
        return send_file(original_path, as_attachment=True)
    else:
        return jsonify({'error': '找不到對應的原始檔案'}), 404


# ------------------------------------------------------------------------------

# 定義檔案列表路由
@app.route('/files', methods=['GET'])
def get_file_list():
    files = os.listdir(UPLOAD_FOLDER)  # 列出上傳資料夾中的檔案
    return jsonify({'files': files}), 200

# ------------------------------------------------------------------------------

# 定義執行動作的路由
@app.route('/perform-action', methods=['POST'])
def perform_action():
    data = request.json  # 取得 JSON 資料
    action = data.get('action')  # 取得動作名稱

    if action == 'start':  # 如果動作是 'start'
        result = "Server received 'start' action and performed the task."
        print(result)
        return jsonify({'status': 'success', 'message': result}), 200
    else:  # 如果是未知動作
        result = f"Server received unknown action: {action}"
        print(result)
        return jsonify({'status': 'error', 'message': 'Unknown action!'}), 400

# ------------------------------------------------------------------------------

def is_flask_running():
    """檢查 Flask (127.0.0.1:5000) 是否已啟動"""
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        s.connect(("127.0.0.1", 5000))
        s.shutdown(socket.SHUT_RDWR)
        return True
    except:
        return False
    finally:
        s.close()

if __name__ == "__main__":
    # 判斷 Flask 是否已經有服務
    if not is_flask_running():
        print("🌐 開啟瀏覽器 http://127.0.0.1:5000")
        webbrowser.open("http://127.0.0.1:5000")
    else:
        print("⚠️ Flask 已在運作，不重複開啟瀏覽器")
    app.run(debug=True, use_reloader=True)











# async def analyze_excel_async(filepath, weights=None):
#     start_time = time.time()
#     default_weights = {
#         'keyword': 5.0,
#         'multi_user': 3.0,
#         'escalation': 2.0,
#         'config_item': 5.0,
#         'role_component': 3.0,
#         'time_cluster': 2.0
#     }
#     weights = {**default_weights, **(weights or {})}

#     df = pd.read_excel(filepath)
#     component_counts = df['Role/Component'].value_counts()
#     configuration_item_counts = df['Configuration item'].value_counts()
#     configuration_item_max = configuration_item_counts.max()
#     df['Opened'] = pd.to_datetime(df['Opened'], errors='coerce')
#     analysis_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

#     # 非同步處理每筆資料
#     tasks = [
#         analyze_row_async(row, idx, df, weights, component_counts, configuration_item_counts, configuration_item_max, analysis_time)
#         for idx, row in df.iterrows()
#     ]
#     results_raw = await asyncio.gather(*tasks)
#     results = [r for r in results_raw if r]

#     # ✅ 分群邏輯（照原本邏輯即可）
#     all_scores = [r['impactScore'] for r in results]
#     score_range = max(all_scores) - min(all_scores)
#     score_std = np.std(all_scores)

#     print(f"📈 分群判斷指標：count={len(all_scores)}, range={score_range:.2f}, stddev={score_std:.2f}")

#     if (
#         len(all_scores) >= KMEANS_MIN_COUNT and
#         score_range >= KMEANS_MIN_RANGE and
#         score_std >= KMEANS_MIN_STDDEV
#     ):
#         kmeans = KMeans(n_clusters=4, random_state=42)
#         labels = kmeans.fit_predict(np.array(all_scores).reshape(-1, 1))
#         centroids = kmeans.cluster_centers_.flatten()
#         set_kmeans_thresholds_from_centroids(centroids)
#         print(f"📊 KMeans 分群標籤：{labels}")
#         label_map = {}
#         for i, idx in enumerate(np.argsort(centroids)[::-1]):
#             label_map[idx] = ['高風險', '中風險', '低風險', '忽略'][i]
#         for i, r in enumerate(results):
#             r['riskLevel'] = label_map[labels[i]]
#         print(f"📌 KMeans 分群中心：{sorted(centroids, reverse=True)}")
#     else:
#         print("⚠️ 不啟用 KMeans，改用固定門檻分級")
#         for r in results:
#             r['riskLevel'] = get_risk_level(r['impactScore'])

#     total_time = time.time() - start_time
#     avg_time = total_time / len(results) if results else 0

#     print(f"\n🎯 所有分析總耗時：{total_time:.2f} 秒")
#     print(f"📊 單筆平均耗時：{avg_time:.2f} 秒")

#     print("\n✅ 所有資料分析完成！")
#     return {
#         'data': results,
#         'analysisTime': analysis_time
#     }







# async def analyze_row_async(row, idx, df, weights, component_counts, configuration_item_counts, configuration_item_max, analysis_time):
#     try:
#         description_text = row.get('Description', 'not filled')
#         short_description_text = row.get('Short description', 'not filled')
#         close_note_text = row.get('Close notes', 'not filled')

#         keyword_score = is_high_risk(short_description_text)
#         user_impact_score = is_multi_user(description_text)
#         escalation_score = is_escalated(close_note_text)

#         config_raw = configuration_item_counts.get(row.get('Configuration item'), 0)
#         configuration_item_freq = config_raw / configuration_item_max if configuration_item_max > 0 else 0

#         role_comp = row.get('Role/Component', 'not filled')
#         count = component_counts.get(role_comp, 0)
#         role_component_freq = 3 if count >= 5 else 2 if count >= 3 else 1 if count == 2 else 0

#         this_time = row.get('Opened', 'not filled')
#         if pd.isnull(this_time):
#             time_cluster_score = 1
#         else:
#             others = df[df['Role/Component'] == role_comp]
#             close_events = others[(others['Opened'] >= this_time - pd.Timedelta(hours=24)) &
#                                   (others['Opened'] <= this_time + pd.Timedelta(hours=24))]
#             count_cluster = len(close_events)
#             time_cluster_score = 3 if count_cluster >= 3 else 2 if count_cluster == 2 else 1

#         severity_score = round(
#             keyword_score * weights['keyword'] +
#             user_impact_score * weights['multi_user'] +
#             escalation_score * weights['escalation'], 2
#         )
#         frequency_score = round(
#             configuration_item_freq * weights['config_item'] +
#             role_component_freq * weights['role_component'] +
#             time_cluster_score * weights['time_cluster'], 2
#         )
#         impact_score = round(math.sqrt(severity_score**2 + frequency_score**2), 2)
#         risk_level = get_risk_level(impact_score)

#         desc = str(description_text).strip()
#         short_desc = str(short_description_text).strip()
#         close_notes = str(close_note_text).strip()
#         resolution_text = f"{desc}\n{short_desc}\n{close_notes}".strip()

#         ai_suggestion, ai_summary = await asyncio.gather(
#             extract_resolution_suggestion(resolution_text),
#             extract_problem_with_custom_prompt(f"{short_desc}\n{desc}".strip())
#         )

#         recommended = recommend_solution(short_description_text)
#         keywords = extract_keywords(short_description_text)

#         return {
#             'id': safe_value(row.get('Incident') or row.get('Number')),
#             'configurationItem': safe_value(row.get('Configuration item')),
#             'roleComponent': safe_value(row.get('Role/Component')),
#             'subcategory': safe_value(row.get('Subcategory')),
#             'aiSummary': safe_value(ai_summary),
#             'originalShortDescription': safe_value(short_desc),
#             'originalDescription': safe_value(desc),
#             'severityScore': safe_value(severity_score),
#             'frequencyScore': safe_value(frequency_score),
#             'impactScore': safe_value(impact_score),
#             'severityScoreNorm': round(severity_score / 10, 2),
#             'frequencyScoreNorm': round(frequency_score / 20, 2),
#             'impactScoreNorm': round(impact_score / 30, 2),
#             'riskLevel': risk_level,
#             'solution': safe_value(ai_suggestion or '無提供解法'),
#             'location': safe_value(row.get('Location')),
#             'analysisTime': analysis_time,
#             'weights': {k: round(v / 10, 2) for k, v in weights.items()},
#         }

#     except Exception as e:
#         print(f"❌ 分析第 {idx+1} 筆失敗：", e)
#         return None











# def analyze_excel(filepath, weights=None):
#         # 預設權重設定（可被覆蓋）
#     default_weights = {
#         'keyword': 5.0,
#         'multi_user': 3.0,
#         'escalation': 2.0,
#         'config_item': 5.0,
#         'role_component': 3.0,
#         'time_cluster': 2.0
#     }
#     weights = {**default_weights, **(weights or {})}  # 合併預設權重與使用者提供的權重設定
#     print("🎛️ 使用中的權重設定：", weights)
#     print("🔍 開始分析 Excel 檔案...")
#     print(f"\n📂 讀取 Excel：{filepath}")
#     df = pd.read_excel(filepath)  # 讀取 Excel 檔案
#     print(f"📊 共讀取 {len(df)} 筆資料\n")
#     component_counts = df['Role/Component'].value_counts()  # 計算每個角色/元件的出現次數
#     df['Opened'] = pd.to_datetime(df['Opened'], errors='coerce')  # 將 'Opened' 欄位轉為日期格式
#     results = []  # 儲存分析結果
#     configuration_item_counts = df['Configuration item'].value_counts()  # 計算每個配置項的出現次數
#     configuration_item_max = configuration_item_counts.max()  # 找出配置項的最大出現次數
#     analysis_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#     print(f"📅 分析時間：{analysis_time}")

#     for idx, row in tqdm(df.iterrows(), total=len(df), desc="📊 分析進度"):
#         print(f"\n🔍 第 {idx + 1} 筆分析中...")
#         description_text = row.get('Description', 'not filled')  # 取得描述文字
#         short_description_text = row.get('Short description', 'not filled') # 取得簡短描述文字
#         close_note_text = row.get('Close notes', 'not filled')  # 取得關閉註解文字
#         print(f"📄 描述：{description_text}")
#         print(f"🔑 簡短描述：{short_description_text}")
#         print(f"🔒 關閉註解：{close_note_text}")
#         # 這裡可以加入對描述文字的預處理，例如去除多餘空格、轉為小寫等
#         # description_text = normalize_text(description_text)  # 標準化文字    


#         #這裡要改成使用語意分析模型

#         keyword_score = is_high_risk(short_description_text)  # 計算關鍵字分數
#         print(f"⚠️ 高風險語意分數（keyword_score）：{keyword_score}")
#         user_impact_score = is_multi_user(description_text)  # 計算使用者影響分數
#         print(f"👥 多人影響分數（user_impact_score）：{user_impact_score}")
#         escalation_score = is_escalated(close_note_text)  # 計算升級處理分數
#         print(f"📈 升級處理分數（escalation_score）：{escalation_score}")



#         config_raw = configuration_item_counts.get(row.get('Configuration item'), 0)  # 取得配置項的出現次數
#         configuration_item_freq = config_raw / configuration_item_max if configuration_item_max > 0 else 0  # 計算配置項頻率

#         role_comp = row.get('Role/Component', 'not filled')  # 取得角色/元件
#         count = component_counts.get(role_comp, 0)  # 取得角色/元件的出現次數
#         if count >= 5:
#             role_component_freq = 3
#         elif count >= 3:
#             role_component_freq = 2
#         elif count == 2:
#             role_component_freq = 1
#         else:
#             role_component_freq = 0

#         this_time = row.get('Opened', 'not filled')  # 取得開啟時間
#         if pd.isnull(this_time):  # 如果開啟時間為空
#             time_cluster_score = 1
#         else:
#             others = df[df['Role/Component'] == role_comp]  # 篩選相同角色/元件的資料
#             close_events = others[(others['Opened'] >= this_time - pd.Timedelta(hours=24)) &
#                                   (others['Opened'] <= this_time + pd.Timedelta(hours=24))]  # 找出 24 小時內的事件
#             count_cluster = len(close_events)  # 計算事件數量
#             if count_cluster >= 3:
#                 time_cluster_score = 3
#             elif count_cluster == 2:
#                 time_cluster_score = 2
#             else:
#                 time_cluster_score = 1

#         severity_score = round(
#             keyword_score * weights['keyword'] +
#             user_impact_score * weights['multi_user'] +
#             escalation_score * weights['escalation'], 2
#         )

#         frequency_score = round(
#             configuration_item_freq * weights['config_item'] +
#             role_component_freq * weights['role_component'] +
#             time_cluster_score * weights['time_cluster'], 2
#         )


        
#         print(f"📊 嚴重性分數：{severity_score}，頻率分數：{frequency_score}")
#         print("🧠 頻率分數細項：")
#         print(f"🔸 配置項（Configuration Item）出現比例：{configuration_item_freq:.2f}，乘以權重後得 {configuration_item_freq * weights['config_item']:.2f} 分")
#         print(f"🔸 元件或角色（Role/Component）在整體中出現 {count} 次 → 給 {role_component_freq * weights['role_component']:.2f} 分")
#         print(f"🔸 在 24 小時內有 {count_cluster} 筆同元件事件 → 群聚加分 {time_cluster_score * weights['time_cluster']:.2f} 分")
#         print(f"📊 頻率總分 = {frequency_score}\n")




#         # 計算影響分數
#         impact_score = round(math.sqrt(severity_score**2 + frequency_score**2), 2)
#         risk_level = get_risk_level(impact_score)
#         print(f"📉 嚴重性：{severity_score}, 頻率：{frequency_score}, 總分(After KMean process)：{impact_score} → 分級：{risk_level}")
#         desc = str(row.get('Description', "")).strip()
#         short_desc = str(row.get('Short Description', "")).strip()
#         close_notes = str(row.get('Close notes', "")).strip()
#         resolution_text = f"{desc}\n{short_desc}\n{close_notes}".strip()

#         print(f"📦 Resolution 原始文字：{resolution_text}")  # ✅ 確認原始欄位內容

#         ai_suggestion = extract_resolution_suggestion(resolution_text)
#         print(f"🤖 GPT 建議句回傳：{ai_suggestion}")  # ✅ 確認 GPT 是否成功回應

#         # ✅ 安全地建立 AI 摘要輸入（若全空則顯示無資料）
#         summary_input_text = f"{short_desc}\n{desc}".strip()
#         if not summary_input_text:
#             summary_input_text = "（無原始摘要輸入）"

#         # ✅ 呼叫 GPT 摘要函式
#         ai_summary = extract_problem_with_custom_prompt(summary_input_text)

#         print(f"📦 Resolution 原始文字：{resolution_text}")
#         print(f"📝 AI 摘要輸入：{summary_input_text}")
#         print(f"🤖 GPT 摘要回傳：{ai_summary}")

#         # 儲存分析結果
#         results.append({
#             'id': safe_value(row.get('Incident') or row.get('Number')),
#             'configurationItem': safe_value(row.get('Configuration item')),
#             'roleComponent': safe_value(row.get('Role/Component')),
#             'subcategory': safe_value(row.get('Subcategory')),
#             'aiSummary': safe_value(ai_summary),
#             'originalShortDescription': safe_value(short_desc),
#             'originalDescription': safe_value(desc),
#             'severityScore': safe_value(severity_score),
#             'frequencyScore': safe_value(frequency_score),
#             'impactScore': safe_value(impact_score),
#             'severityScoreNorm': round(severity_score / 10, 2),
#             'frequencyScoreNorm': round(frequency_score / 20, 2),
#             'impactScoreNorm': round(impact_score / 30, 2),
#             'riskLevel': safe_value(get_risk_level(impact_score)),
#             'solution': safe_value(ai_suggestion or '無提供解法'),
#             'location': safe_value(row.get('Location')),
#             'analysisTime': analysis_time,
#             'weights': {k: round(v / 10, 2) for k, v in weights.items()},
#         })

#         # solution_text = row.get('Close notes') or '無提供解法'
#         recommended = recommend_solution(short_description_text)
#         keywords = extract_keywords(short_description_text)

#         print(f"✅ 已儲存 solution：{results[-1]['solution']}")
#         print(f"💡 建議解法：{recommended}")
#         print(f"🔑 抽取關鍵字：{keywords}")
#         print("—" * 250)  # 分隔線




#     # ⬇⬇⬇ KMeans 分群邏輯（支援三條件） ⬇⬇⬇
#     all_scores = [r['impactScore'] for r in results]
#     score_range = max(all_scores) - min(all_scores)
#     score_std = np.std(all_scores)

#     print(f"📈 分群判斷指標：count={len(all_scores)}, range={score_range:.2f}, stddev={score_std:.2f}")

#     if (
#         len(all_scores) >= KMEANS_MIN_COUNT and
#         score_range >= KMEANS_MIN_RANGE and
#         score_std >= KMEANS_MIN_STDDEV
#     ):
#         kmeans = KMeans(n_clusters=4, random_state=42)
#         labels = kmeans.fit_predict(np.array(all_scores).reshape(-1, 1))
#         centroids = kmeans.cluster_centers_.flatten()
#         set_kmeans_thresholds_from_centroids(centroids)
#         print(f"📊 KMeans 分群標籤：{labels}")
#         label_map = {}
#         for i, idx in enumerate(np.argsort(centroids)[::-1]):
#             label_map[idx] = ['高風險', '中風險', '低風險', '忽略'][i]
#         for i, r in enumerate(results):
#             r['riskLevel'] = label_map[labels[i]]
#         print(f"📌 KMeans 分群中心：{sorted(centroids, reverse=True)}")
#     else:
#         print("⚠️ 不啟用 KMeans，改用固定門檻分級")
#         for r in results:
#             r['riskLevel'] = get_risk_level(r['impactScore'])
#     # ⬆⬆⬆ 分群邏輯結束 ⬆⬆⬆
#     print("\n✅ 所有資料分析完成！")
#     return {
#         'data': results,
#         'analysisTime': analysis_time
#     }








